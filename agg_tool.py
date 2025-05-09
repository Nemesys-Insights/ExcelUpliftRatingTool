import argparse
import os
import json
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from Metrics import MetricsDictionary


# === CONFIGURATION ===
REQUIRED_COLUMNS = ["ReviewerID",
                    "UID",
                    "Architecture ID",
                    "Reason 1",
                    "Reason 2",
                    "Reason 3",
                    "Reason 4",
                    "Reason 5",
                    "Category",
                    "Element",
                    "Task",
                    "Prompt",
                    "Response",
                    "Accuracy",
                    "Completeness",
                    "Novelty",
                    "Likelihood of Acceptance",
                    "Safety"]  
VALID_VALUES = MetricsDictionary

FILL_INVALID = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")


def remove_no_break_space(text):
    if text is not pd.NA:
      return text.replace('\xa0', '')
    return pd.NA
    
def load_excel(file_path):
    """Load an excel workbook and return a tuple containing the workbook, the active worksheet, and a dataframe created from the excel."""

    wb = load_workbook(file_path)
    ws = wb.active
    df :pd.DataFrame = pd.read_excel(file_path,
                       sheet_name="Responses",
                       skiprows=4, #Skip header content rows
                       usecols="B:P,R:W", #Skip over spacer columns that are only present for readability
                       header= 0
                       )
  
    for col in VALID_VALUES.keys():
        df[col] = df[col].astype('string')
    return df, wb, ws


def check_required_columns(df) -> list:
    """Confirm that the required columns are present in the dataframe.
    
    Returns a list of issues. The list is empty if there are no issues."""

    issues = []
    required = REQUIRED_COLUMNS + list(VALID_VALUES.keys())
    for col in required:
        if col not in df.columns:
            issues.append({"error": "Missing required column", "column": col})
    return issues


def check_required_values(df):
    """Confirm that the each row of the dataframe has a value for every required column."""

    issues = []
    df["Missing Value"] = False
    for row_idx, row in df.iterrows():
        for col in REQUIRED_COLUMNS:
            if pd.isna(row[col]):
                issues.append({
                    "row": int(row_idx + 2),
                    "column": col,
                    "error": "Missing required value"
                })
                row["Missing Value"] = True
                df.iloc[row_idx] = row
            
    return issues


def check_dropdowns(df, valid_values):
    """Confirm that the values expected for a given column are valid based on the 
    schema defined in valid_values."""

    issues = []
    df["Invalid Dropdown Value"] = False
    for row_idx, row in df.iterrows():
        for col in valid_values:
            if col in df.columns:
                val = row[col]
                if str(val) not in valid_values[col] and not pd.isna(val):
                    issues.append({
                        "row": int(row_idx + 2),
                        "column": col,
                        "value": str(val),
                        "error": "Invalid dropdown value"
                    })
                    row["Invalid Dropdown Value"] = True
                    df.iloc[row_idx] = row
    return issues




def append_issues_sheet(wb, issues):
    """Adds a sheet that lists all issues with the input sheet."""

    ws_issues = wb.create_sheet(title="Validation Issues")
    if issues:
        headers = sorted(issues[0].keys())
        ws_issues.append(headers)
        for issue in issues:
            ws_issues.append([issue.get(h, "") for h in headers])
    else:
        ws_issues.append(["No validation issues found."])


def validate_excel(file_path):
    """Check for issues with the input data in the specified excel file, highlight cells with issues and append a list of issues 
    as a new sheet in the workbook."""
    df, wb, ws = load_excel(file_path)

    issues = []
    issues += check_required_columns(df)

    if df.empty:
        issues.append({"error": "No data rows found in the sheet."})

    column_issues = issues.copy()
    issues += check_required_values(df)
    issues += check_dropdowns(df, VALID_VALUES)
    append_issues_sheet(wb, issues)
    
    if column_issues:
        return issues, None, wb
    return issues, df, wb
   


def agg_data(input_directory):



    input_dir = Path(input_directory)
    if not input_dir.exists() or not input_dir.is_dir():
        raise ValueError("Input path is not a valid directory.")

    #name the output directory with a timestamp for uniqueness
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = input_dir.parent / f"output_{timestamp}"
    annotated_dir = output_dir / "Annotated Inputs"
    output_dir.mkdir()
    annotated_dir.mkdir()

    validation_log = []
    all_data = []
    reviewer_ids_seen = set()
    duplicate_reviewers = set()

    for file in input_dir.glob("*.xlsx"):
        print(f"Processing {file.name}...")
        issues, df, wb = validate_excel(file)

        if issues:
            print(f"  Found {len(issues)} issue(s)")
            
        print("  No issues found")
        if df is not None:
            reviewer_ids = df["ReviewerID"].unique()
            for reviewer_id in reviewer_ids:
                if reviewer_id in reviewer_ids_seen:
                    print(f"  WARNING: Duplicate Reviewer ID detected: {reviewer_id}")
                    duplicate_reviewers.add(reviewer_id)
                reviewer_ids_seen.add(reviewer_id)
            all_data.append(df)

        annotated_path = annotated_dir / f"annotated_{file.name}"
        wb.save(annotated_path)

        validation_log.append({"file": file.name, "issues": issues})

    validation_df = pd.DataFrame([
        {"file": entry["file"], **issue}
        for entry in validation_log for issue in entry["issues"]
    ])
    if validation_df.empty:
        validation_df = pd.DataFrame([{"file": "ALL", "status": "No validation issues found"}])
    validation_df.to_excel(output_dir / "validation_log.xlsx", index=False)

    
    if duplicate_reviewers:
        dup_rows = []
        for file in input_dir.glob("*.xlsx"):
            df, _, _ = load_excel(file)
            if "Reviewer ID" in df.columns:
                dup_df = df[df["Reviewer ID"].isin(duplicate_reviewers)]
                for reviewer_id in dup_df["Reviewer ID"].unique():
                    reviewer_rows = dup_df[dup_df["Reviewer ID"] == reviewer_id]
                    for row_id in reviewer_rows["Row ID"].unique():
                        dup_rows.append({
                            "Reviewer ID": reviewer_id,
                            "Row ID": row_id,
                            "File": file.name
                        })

        dup_df_expanded = pd.DataFrame(dup_rows)
        with pd.ExcelWriter(output_dir / "validation_log.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            dup_df_expanded.to_excel(writer, sheet_name="Duplicate Reviewers", index=False)


    if all_data:
        combined = pd.concat(all_data, ignore_index=True)
        combined.to_excel(output_dir / "combined_clean_data.xlsx", index=False)

        print("Aggregated clean data saved.")
    else:
        print("No clean data to aggregate.")

    return f"Results saved to: {output_dir}"