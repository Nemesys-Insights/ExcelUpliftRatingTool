import os
import shutil
import random
import argparse
import sys 
from pathlib import Path
import math 

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def resource_path(rel_path: str | Path) -> Path:
    """
    Return an absolute Path to a bundled resource, whether the code is running
    from source or from a PyInstaller one‑file executable.
    """
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
    return base / rel_path

# Path to your formatted template
TEMPLATE_PATH = resource_path('template/eval_template.xlsx')
# Table starts at row 6, headers are in row 5
TABLE_START_ROW = 6


def assign_rows(df : pd.DataFrame, num_evaluators : int, evaluators_per_row : int = 3):
    """
    """
    print(f"Evaluators: {num_evaluators}")
    # initialize assignment lists and load counts
    assignments = {i: [] for i in range(1, num_evaluators + 1)}
    load = {i: 0 for i in assignments}

    
    print(f"reviewers per row: {evaluators_per_row}")
    for _, row in df.iterrows():
        # find evaluators who haven't reviewed this Response ID
        eligible = [
            ev for ev, rows in assignments.items()
            if all(not (r_assigned['UID'] == row['UID']) for r_assigned in rows)
        ]

        if len(eligible) < evaluators_per_row:
            raise ValueError(
                f"Cannot find {evaluators_per_row} fresh evaluators for UID {row['UID']}.\
                Only {len(eligible)} fresh reviewers available."
            )

        # sort by current load (ascending); break ties by shuffling
        # group by load value
        random.shuffle(eligible)
        eligible.sort(key=lambda ev: load[ev])

        # pick the three least-loaded
        chosen = eligible[:evaluators_per_row]

        # record assignments and bump their load
        for ev in chosen:
            assignments[ev].append(row.to_dict())
            load[ev] += 1

    return assignments


def output_from_template(assignments, output_folder):
    """
    For each evaluator:
      1. Copy the formatted template to evaluator_{n}.xlsx
      2. Clear existing data
      3. Fill in assigned rows starting at row 6,
      4. Lock data columns for preservation
    """
    os.makedirs(output_folder, exist_ok=True)
    from openpyxl.styles import Protection
    for evaluator, rows in assignments.items():
        dest = os.path.join(output_folder, f"/Evaluator {evaluator}/evaluator_{evaluator}.xlsx")
        shutil.copy(TEMPLATE_PATH, dest)
        wb = load_workbook(dest)
        ws = wb.active

        # Clear old data in B6:J...
        for r in ws.iter_rows(min_row=TABLE_START_ROW, min_col=2, max_col=10, max_row=ws.max_row):
            for cell in r:
                cell.value = None

        columns = [
            'UID',
            'Architecture ID',
            'Batch ID', 
            'Jailbroken Prompt',
            'Reason 1', 
            'Reason 2', 
            'Reason 3', 
            'Reason 4' ,
            'Reason 5' , 
            'Category',
            'Element',
            'Task',
            'Prompt',
            'Response'
            ] 
        
        # Populate data
        rowcount = 0
        for i, r in enumerate(rows):
            row_idx = TABLE_START_ROW + i
            ws.cell(row_idx, 2, evaluator)  # B: ReviewerID
            rowcount = 2

            #insert the rest of the columns
            for idx, col in enumerate(columns):
                rowcount = 2+idx+1
                ws.cell(row_idx,rowcount, r[col])
                
            ws.cell(row_idx, rowcount+1, ".") #F : Separator (to prevent overflow)
           

        # Lock base cols and unlock grading cols
        for i in range(len(rows)):
            row_idx = TABLE_START_ROW + i
            for col in range(2, rowcount):  # lock all data columns
                ws.cell(row_idx, col).protection = Protection(locked=True)
            for col in range(rowcount, rowcount+8): # F-N unlocked
                ws.cell(row_idx, col).protection = Protection(locked=False)

        ws.protection.sheet = True
        ws.views.sheetView[0].selection[0].sqref = "A1"
        wb.save(dest)
        print(f"Created {dest} with {len(rows)} rows.")


def output_mapping_workbook(assignments, output_folder):
    """
    Writes assignment_mapping.xlsx with two sheets:
      - 'Assignments' starting at row 6, cols B-F (Architecture, Response ID, Prompt, Response, Reviewers)
      - 'Summary' listing ReviewerID and count of assigned prompts.
    """
    mapping = {}
    for ev, rows in assignments.items():
        for r in rows:
            uid = r['UID']
            if uid not in mapping:
                mapping[uid] = {
                    'UID': uid,
                    'Architecture ID': r['Architecture ID'],
                    'Batch ID': r['Batch ID'],
                    'Prompt': r['Prompt'],
                    'Response': r['Response'],
                    'Reviewers': []
                }
            mapping[uid]['Reviewers'].append(ev)
  
    df_map = pd.DataFrame([{
        'UID': v['UID'],
        'Architecture ID': v['Architecture ID'],
        'Batch ID': v['Batch ID'],
        'Prompt': v['Prompt'],
        'Response': v['Response'],
        'Reviewers': ', '.join(map(str, sorted(v['Reviewers'])))
    } for v in mapping.values()])


    path = os.path.join(output_folder, 'assignment_mapping.xlsx')
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # Assignments sheet
        df_map.to_excel(writer, index=False, startrow=1, sheet_name='Assignments', header=True)
        # write headers manually at row 5, cols B-F
        wb = writer.book
        ws = wb['Assignments']

        # Summary sheet
        summary = pd.DataFrame([{'ReviewerID': ev, 'Count': len(rows)}
                                 for ev, rows in assignments.items()])
        summary.to_excel(writer, index=False, startrow=0, sheet_name='Summary')
    print(f"Created companion mapping workbook: {path}")


def assign_workbooks(input_file, output_folder, num_evaluators, evaluators_per_row : int = 3):
    
    if num_evaluators < 3 or num_evaluators > 20:
        raise SystemExit('num_evaluators must be 3–20')
    df = pd.read_excel(input_file)
    for c in ['UID', 'Architecture ID','Jailbroken Prompt','Reason 1', 'Reason 2', 'Reason 3', 'Reason 4' ,'Reason 5', 'Category','Element','Task','Batch ID', 'Prompt','Response']:
        if c not in df.columns:
            raise SystemExit(f'Missing column {c}')

    assignments = assign_rows(df, num_evaluators, evaluators_per_row)
    output_from_template(assignments, output_folder)
    output_mapping_workbook(assignments, output_folder)
    print('Done.')
